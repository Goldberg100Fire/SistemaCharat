from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import firebase_admin
from firebase_admin import credentials, firestore, storage 
from PIL import Image 
import io
import calendar
from datetime import datetime, timedelta
from collections import Counter
import re
import uuid 
import os
import json
import base64
import threading
from werkzeug.security import generate_password_hash, check_password_hash
from google.cloud.firestore_v1 import ArrayUnion
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from gmail_api import send_email

# 1. Configuración de Firebase (local: credenciales.json | Render: FIREBASE_CREDENTIALS_B64)
def _decode_service_account_b64(b64: str) -> dict:
    """Acepta base64 pegado desde el portapapeles (espacios/saltos de línea)."""
    raw = "".join(b64.split())
    pad = (-len(raw)) % 4
    if pad:
        raw += "=" * pad
    decoded = base64.b64decode(raw)
    return json.loads(decoded.decode("utf-8"))


def _firebase_certificate():
    cred_path = "credenciales.json"
    if os.path.isfile(cred_path):
        return credentials.Certificate(cred_path)
    b64 = os.environ.get("FIREBASE_CREDENTIALS_B64")
    if b64:
        try:
            data = _decode_service_account_b64(b64)
            return credentials.Certificate(data)
        except (json.JSONDecodeError, ValueError, UnicodeDecodeError) as e:
            raise RuntimeError(
                "Firebase: FIREBASE_CREDENTIALS_B64 no es un JSON válido en base64. "
                "Vuelve a generarlo en PowerShell y pégalo en Render en una sola línea, sin comillas."
            ) from e
    raise RuntimeError(
        "Firebase: crea credenciales.json en la raíz o define la variable de entorno FIREBASE_CREDENTIALS_B64."
    )

if not firebase_admin._apps:
    cred = _firebase_certificate()
    firebase_admin.initialize_app(cred, {
        'storageBucket': 'gestion-charat-admin.firebasestorage.app' 
    })

db = firestore.client()
bucket = storage.bucket() 

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "muni_charat_2026_secure_key")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(BASE_DIR, "static", "img", "logo_charat.png")

CATEGORIAS_INICIALES = (
    "Agua y Saneamiento",
    "Vías y Caminos",
    "Alumbrado Público",
    "Limpieza Pública",
)

ICONOS_CATEGORIA = {
    "Agua y Saneamiento": "fa-droplet",
    "Vías y Caminos": "fa-road",
    "Alumbrado Público": "fa-lightbulb",
    "Limpieza Pública": "fa-broom",
}

SLA_DIAS_CATEGORIA = {
    "Agua y Saneamiento": 3,
    "Vías y Caminos": 15,
    "Alumbrado Público": 7,
    "Limpieza Pública": 5,
}
MESES_CORTO = ("Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic")
DIAS_SEMANA = ("Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom")


def _admin_redirect(mensaje=None, categoria="success"):
    if mensaje:
        flash(mensaje, categoria)
    return redirect(url_for("admin"))


def _generar_folio():
    return f"CHR-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"


def _sembrar_categorias_iniciales():
    try:
        if not db.collection("categorias").limit(1).get():
            ahora = datetime.now()
            for nombre in CATEGORIAS_INICIALES:
                db.collection("categorias").add({"nombre": nombre, "fecha_registro": ahora})
            print("LOG: Categorías iniciales creadas en Firebase.")
    except Exception as e:
        print(f"LOG: No se pudieron sembrar categorías: {e}")


def _icono_categoria(nombre):
    return ICONOS_CATEGORIA.get(nombre, "fa-tag")


def _estadisticas_reportes(reportes):
    conteo = {"total": len(reportes), "pendiente": 0, "en_proceso": 0, "atendido": 0}
    for r in reportes:
        estado = r.get("gestion", {}).get("estado", "Pendiente")
        if estado == "Pendiente":
            conteo["pendiente"] += 1
        elif estado == "En Proceso":
            conteo["en_proceso"] += 1
        elif estado == "Atendido":
            conteo["atendido"] += 1
    return conteo


def _fecha_valor(val):
    if val is None:
        return None
    if hasattr(val, "timestamp"):
        dt = val
        if getattr(dt, "tzinfo", None) is not None:
            dt = dt.replace(tzinfo=None)
        return dt
    return None


def _dias_transcurridos(inicio, fin=None):
    if not inicio:
        return 0
    fin = fin or datetime.now()
    return max(0, (fin - inicio).total_seconds() / 86400)


def _estadisticas_avanzadas(reportes):
    ahora = datetime.now()
    hoy = ahora.date()
    inicio_mes = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if inicio_mes.month == 1:
        inicio_mes_anterior = inicio_mes.replace(year=inicio_mes.year - 1, month=12)
    else:
        inicio_mes_anterior = inicio_mes.replace(month=inicio_mes.month - 1)

    por_categoria = Counter()
    por_caserio = Counter()
    mes_actual = mes_anterior = hoy_count = con_gps = 0
    ultimos_7 = {(hoy - timedelta(days=i)).strftime("%d/%m"): 0 for i in range(6, -1, -1)}
    tiempos_resolucion_horas = []
    por_dia_semana = {i: 0 for i in range(7)}
    por_hora = {h: 0 for h in range(24)}
    operadores = Counter()
    sla_stats = {}
    meses_hist = {}

    for i in range(5, -1, -1):
        m = ahora.month - i
        y = ahora.year
        while m <= 0:
            m += 12
            y -= 1
        meses_hist[(y, m)] = 0

    for r in reportes:
        cat = r.get("detalle", {}).get("categoria", "Sin categoría")
        por_categoria[cat] += 1
        por_caserio[r.get("ubicacion", {}).get("caserio", "Sin sector")] += 1
        if r.get("ubicacion", {}).get("lat"):
            con_gps += 1

        fr = _fecha_valor(r.get("gestion", {}).get("fecha_registro"))
        if fr:
            fd = fr.date()
            clave = fd.strftime("%d/%m")
            if clave in ultimos_7:
                ultimos_7[clave] += 1
            if fd == hoy:
                hoy_count += 1
            if fr >= inicio_mes:
                mes_actual += 1
            elif inicio_mes_anterior <= fr < inicio_mes:
                mes_anterior += 1
            por_dia_semana[fr.weekday()] += 1
            por_hora[fr.hour] += 1
            key_m = (fr.year, fr.month)
            if key_m in meses_hist:
                meses_hist[key_m] += 1

        for h in r.get("gestion", {}).get("historial") or []:
            u = h.get("usuario", "")
            if u and u not in ("Ciudadano", "Sistema") and h.get("estado") in ("En Proceso", "Atendido"):
                operadores[u] += 1

        estado = r.get("gestion", {}).get("estado", "Pendiente")
        sla_dias = SLA_DIAS_CATEGORIA.get(cat, 7)
        if cat not in sla_stats:
            sla_stats[cat] = {"categoria": cat, "sla_dias": sla_dias, "total": 0, "fuera_plazo": 0, "atendidas": 0, "en_plazo": 0}
        sla_stats[cat]["total"] += 1

        if estado == "Atendido" and fr:
            fin = fr
            for h in r.get("gestion", {}).get("historial") or []:
                if h.get("estado") == "Atendido":
                    hf = _fecha_valor(h.get("fecha"))
                    if hf:
                        fin = hf
            dias = _dias_transcurridos(fr, fin)
            sla_stats[cat]["atendidas"] += 1
            if dias <= sla_dias:
                sla_stats[cat]["en_plazo"] += 1
            horas = (fin - fr).total_seconds() / 3600
            if horas >= 0:
                tiempos_resolucion_horas.append(horas)
        elif fr and _dias_transcurridos(fr) > sla_dias:
            sla_stats[cat]["fuera_plazo"] += 1

    total = len(reportes)
    atendidos = sum(1 for r in reportes if r.get("gestion", {}).get("estado") == "Atendido")
    tasa = round(atendidos / total * 100, 1) if total else 0
    tiempo_prom = round(sum(tiempos_resolucion_horas) / len(tiempos_resolucion_horas) / 24, 1) if tiempos_resolucion_horas else None

    variacion_mes = None
    if mes_anterior:
        variacion_mes = round((mes_actual - mes_anterior) / mes_anterior * 100, 1)

    dias_mes = calendar.monthrange(ahora.year, ahora.month)[1]
    proyeccion_mes = round(mes_actual / max(ahora.day, 1) * dias_mes) if mes_actual else 0

    sla_lista = []
    for item in sla_stats.values():
        pct = round(item["en_plazo"] / item["atendidas"] * 100, 1) if item["atendidas"] else None
        item["cumplimiento_pct"] = pct
        item["alerta"] = item["fuera_plazo"] > 0
        sla_lista.append(item)
    sla_lista.sort(key=lambda x: x["fuera_plazo"], reverse=True)

    return {
        "por_categoria": [{"nombre": k, "total": v} for k, v in por_categoria.most_common()],
        "por_caserio": [{"nombre": k, "total": v} for k, v in por_caserio.most_common()],
        "ultimos_7_dias": [{"fecha": k, "total": v} for k, v in ultimos_7.items()],
        "por_mes": [{"mes": f"{MESES_CORTO[m - 1]} {y}", "total": c} for (y, m), c in sorted(meses_hist.items())],
        "por_dia_semana": [{"dia": DIAS_SEMANA[i], "total": por_dia_semana[i]} for i in range(7)],
        "por_hora": [{"hora": f"{h:02d}:00", "total": por_hora[h]} for h in range(24)],
        "ranking_operadores": [{"usuario": u, "acciones": n} for u, n in operadores.most_common(8)],
        "sla_categorias": sla_lista,
        "proyeccion_mes": proyeccion_mes,
        "tasa_resolucion": tasa,
        "mes_actual": mes_actual,
        "mes_anterior": mes_anterior,
        "variacion_mes_pct": variacion_mes,
        "hoy": hoy_count,
        "con_gps": con_gps,
        "tiempo_promedio_dias": tiempo_prom,
    }


def _password_hasheada(valor):
    return valor.startswith("pbkdf2:") or valor.startswith("scrypt:")


def _verificar_password(almacenada, ingresada):
    if not almacenada:
        return False
    if _password_hasheada(almacenada):
        return check_password_hash(almacenada, ingresada)
    return almacenada == ingresada


def _format_fecha(val):
    if hasattr(val, "strftime"):
        return val.strftime("%d/%m/%Y %H:%M")
    return str(val) if val else "—"


def _obtener_reportes():
    reportes = [doc.to_dict() | {"id": doc.id} for doc in db.collection("incidencias").stream()]
    reportes.sort(key=lambda x: x["gestion"].get("fecha_registro", datetime.now()), reverse=True)
    return reportes


def _filtrar_reportes(reportes, args):
    estado = args.get("estado", "todos")
    caserio = args.get("caserio", "").strip()
    if estado and estado != "todos":
        reportes = [r for r in reportes if r.get("gestion", {}).get("estado") == estado]
    if caserio:
        reportes = [r for r in reportes if r.get("ubicacion", {}).get("caserio") == caserio]
    return reportes


def _puntos_mapa(reportes):
    puntos = []
    for r in reportes:
        lat = r.get("ubicacion", {}).get("lat")
        lng = r.get("ubicacion", {}).get("lng")
        if not lat or not lng:
            continue
        try:
            puntos.append({
                "lat": float(lat),
                "lng": float(lng),
                "folio": r.get("gestion", {}).get("folio", ""),
                "estado": r.get("gestion", {}).get("estado", ""),
                "categoria": r.get("detalle", {}).get("categoria", ""),
                "caserio": r.get("ubicacion", {}).get("caserio", ""),
            })
        except (TypeError, ValueError):
            continue
    return puntos


def _registrar_historial(ref, estado, usuario):
    entrada = {"estado": estado, "usuario": usuario, "fecha": datetime.now()}
    ref.update({
        "gestion.estado": estado,
        "gestion.ultima_actualizacion": datetime.now(),
        "gestion.historial": ArrayUnion([entrada]),
    })


def _eliminar_documentos_coleccion(nombre_coleccion):
    docs = list(db.collection(nombre_coleccion).stream())
    eliminados = 0
    batch = db.batch()
    for doc in docs:
        batch.delete(doc.reference)
        eliminados += 1
        if eliminados % 400 == 0:
            batch.commit()
            batch = db.batch()
    if eliminados % 400 != 0:
        batch.commit()
    return eliminados


def _eliminar_fotos_incidencias():
    eliminadas = 0
    for blob in bucket.list_blobs(prefix="incidencias/"):
        blob.delete()
        eliminadas += 1
    return eliminadas


def _limpiar_datos_prueba():
    incidencias = _eliminar_documentos_coleccion("incidencias")
    fotos = _eliminar_fotos_incidencias()
    return {"incidencias": incidencias, "fotos": fotos}


COLOR_VERDE = "16A34A"
COLOR_VERDE_OSC = "15803D"
COLOR_AMARILLO = "FACC15"
COLOR_AZUL = "2563EB"
COLOR_GRIS = "F8FAFC"
COLOR_GRIS_BORDE = "CBD5E1"
COLOR_TEXTO = "1E293B"
COLOR_TEXTO_SUAVE = "64748B"

ESTADO_FILL_EXCEL = {
    "Pendiente": "FEF9C3",
    "En Proceso": "DBEAFE",
    "Atendido": "DCFCE7",
}

COLUMNAS_EXPORT = [
    "Folio", "Fecha registro", "Informante", "DNI", "Caserío",
    "Categoría", "Estado", "Coordenadas GPS", "Descripción",
]


def _etiqueta_filtros_export(args):
    estado = args.get("estado", "todos")
    caserio = (args.get("caserio") or "").strip()
    partes = [f"Estado: {estado if estado != 'todos' else 'Todos'}"]
    partes.append(f"Caserío: {caserio if caserio else 'Todos'}")
    return " · ".join(partes)


def _resumen_estados(reportes):
    resumen = {"Pendiente": 0, "En Proceso": 0, "Atendido": 0}
    for r in reportes:
        est = r.get("gestion", {}).get("estado", "Pendiente")
        if est in resumen:
            resumen[est] += 1
    return resumen


def _format_gps(r):
    lat = r.get("ubicacion", {}).get("lat")
    lng = r.get("ubicacion", {}).get("lng")
    if lat and lng:
        try:
            return f"{float(lat):.5f}, {float(lng):.5f}"
        except (TypeError, ValueError):
            pass
    return "—"


def _excel_borde():
    thin = Side(style="thin", color=COLOR_GRIS_BORDE)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _excel_bandas_institucionales(ws, col_fin):
    bandas = [(COLOR_AMARILLO, 4), (COLOR_VERDE, 5), (COLOR_AZUL, 6)]
    for color, row in bandas:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_fin)
        cell = ws.cell(row, 1)
        cell.fill = PatternFill("solid", fgColor=color)
        ws.row_dimensions[row].height = 5


def _excel_insertar_logo(ws, celda="A1", ancho=72, alto=72):
    if not os.path.isfile(LOGO_PATH):
        return
    img = XLImage(LOGO_PATH)
    img.width = ancho
    img.height = alto
    ws.add_image(img, celda)


def _excel_membrete(ws, col_fin, titulo, subtitulo):
    _excel_insertar_logo(ws)
    for row, texto, size, bold, color in [
        (1, titulo, 15, True, COLOR_VERDE_OSC),
        (2, "Provincia de Pataz — Departamento de La Libertad, Perú", 9, False, COLOR_TEXTO_SUAVE),
        (3, subtitulo, 10, False, COLOR_TEXTO),
    ]:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=col_fin)
        cell = ws.cell(row, 2, texto)
        cell.font = Font(bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for row in range(1, 4):
        ws.row_dimensions[row].height = 22
    ws.column_dimensions["A"].width = 12
    _excel_bandas_institucionales(ws, col_fin)


def _excel_estilo_celda(cell, *, fill=None, bold=False, color=COLOR_TEXTO, align="left", wrap=False):
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=bold, size=10, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _excel_borde()


def _exportar_excel(reportes, filtros=None):
    filtros = filtros or {}
    wb = Workbook()
    borde = _excel_borde()
    generado = datetime.now().strftime("%d/%m/%Y %H:%M")
    resumen = _resumen_estados(reportes)
    n_cols = len(COLUMNAS_EXPORT)
    col_fin = n_cols

    ws_res = wb.active
    ws_res.title = "Resumen"
    ws_res.sheet_properties.tabColor = COLOR_VERDE
    _excel_membrete(ws_res, 4, "MUNICIPALIDAD DISTRITAL DE CHARAT", "Reporte oficial de incidencias ciudadanas")

    ws_res.merge_cells("A8:D8")
    ws_res["A8"] = "INFORMACIÓN DEL REPORTE"
    ws_res["A8"].font = Font(bold=True, size=11, color="FFFFFF")
    ws_res["A8"].fill = PatternFill("solid", fgColor=COLOR_VERDE_OSC)
    ws_res["A8"].alignment = Alignment(horizontal="center", vertical="center")
    ws_res.row_dimensions[8].height = 24

    meta = [
        ("Fecha y hora de generación", generado),
        ("Filtros aplicados", _etiqueta_filtros_export(filtros)),
        ("Total de registros exportados", len(reportes)),
        ("Registros con coordenadas GPS", sum(1 for r in reportes if _format_gps(r) != "—")),
    ]
    for i, (etiq, valor) in enumerate(meta, start=9):
        ws_res.cell(i, 1, etiq)
        ws_res.cell(i, 2, valor)
        _excel_estilo_celda(ws_res.cell(i, 1), bold=True, color=COLOR_TEXTO_SUAVE)
        _excel_estilo_celda(ws_res.cell(i, 2), bold=True, color=COLOR_TEXTO)
        ws_res.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

    ws_res.merge_cells("A14:D14")
    ws_res["A14"] = "INDICADORES POR ESTADO"
    ws_res["A14"].font = Font(bold=True, size=11, color="FFFFFF")
    ws_res["A14"].fill = PatternFill("solid", fgColor=COLOR_VERDE_OSC)
    ws_res["A14"].alignment = Alignment(horizontal="center", vertical="center")

    kpi_cfg = [
        ("TOTAL", len(reportes), COLOR_VERDE, "FFFFFF"),
        ("PENDIENTES", resumen["Pendiente"], "FEF9C3", COLOR_TEXTO),
        ("EN PROCESO", resumen["En Proceso"], "DBEAFE", COLOR_TEXTO),
        ("ATENDIDAS", resumen["Atendido"], "DCFCE7", COLOR_TEXTO),
    ]
    for col, (lbl, val, fondo, texto) in enumerate(kpi_cfg, start=1):
        c_lbl = ws_res.cell(15, col, lbl)
        c_val = ws_res.cell(16, col, val)
        _excel_estilo_celda(c_lbl, fill=fondo, bold=True, color=texto, align="center")
        _excel_estilo_celda(c_val, fill=fondo, bold=True, color=texto, align="center")
        c_val.font = Font(bold=True, size=18, color=texto)
        ws_res.column_dimensions[get_column_letter(col)].width = 18

    ws_res.merge_cells("A18:D18")
    ws_res["A18"] = "Documento generado automáticamente por el Sistema MuniCharat. Uso institucional."
    ws_res["A18"].font = Font(italic=True, size=9, color=COLOR_TEXTO_SUAVE)
    ws_res["A18"].alignment = Alignment(horizontal="center")

    ws = wb.create_sheet("Incidencias")
    ws.sheet_properties.tabColor = COLOR_AZUL
    _excel_membrete(ws, col_fin, "MUNICIPALIDAD DISTRITAL DE CHARAT", "Detalle de incidencias registradas")
    fila_meta = 8
    ws.merge_cells(start_row=fila_meta, start_column=1, end_row=fila_meta, end_column=col_fin)
    meta_cell = ws.cell(fila_meta, 1, f"Generado: {generado}   |   {_etiqueta_filtros_export(filtros)}   |   Total: {len(reportes)}")
    meta_cell.font = Font(size=9, color=COLOR_TEXTO_SUAVE)
    meta_cell.alignment = Alignment(horizontal="center", vertical="center")
    meta_cell.fill = PatternFill("solid", fgColor=COLOR_GRIS)
    meta_cell.border = borde
    ws.row_dimensions[fila_meta].height = 20

    fila_inicio = 10
    header_fill = PatternFill("solid", fgColor=COLOR_VERDE_OSC)
    for col, titulo in enumerate(COLUMNAS_EXPORT, 1):
        cell = ws.cell(fila_inicio, col, titulo)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde

    for i, r in enumerate(reportes):
        fila = fila_inicio + 1 + i
        estado = r.get("gestion", {}).get("estado", "")
        valores = [
            r.get("gestion", {}).get("folio", ""),
            _format_fecha(r.get("gestion", {}).get("fecha_registro")),
            r.get("informante", {}).get("nombre", ""),
            r.get("informante", {}).get("dni", ""),
            r.get("ubicacion", {}).get("caserio", ""),
            r.get("detalle", {}).get("categoria", ""),
            estado,
            _format_gps(r),
            r.get("detalle", {}).get("descripcion", ""),
        ]
        fondo_fila = COLOR_GRIS if i % 2 else "FFFFFF"
        for col, val in enumerate(valores, 1):
            cell = ws.cell(fila, col, val)
            fondo = ESTADO_FILL_EXCEL.get(estado, fondo_fila) if col == 7 else fondo_fila
            cell.fill = PatternFill("solid", fgColor=fondo)
            cell.border = borde
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (8, 9)))
            if col == 1:
                cell.font = Font(bold=True, color=COLOR_VERDE_OSC, size=10)
            elif col == 7:
                cell.font = Font(bold=True, size=9, color=COLOR_TEXTO)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = Font(size=10, color=COLOR_TEXTO)

    anchos = [20, 18, 24, 12, 18, 22, 14, 20, 48]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ultima_fila = fila_inicio + len(reportes)
    ws.freeze_panes = ws.cell(fila_inicio + 1, 1)
    ws.row_dimensions[fila_inicio].height = 28
    ws.auto_filter.ref = f"A{fila_inicio}:{get_column_letter(col_fin)}{max(fila_inicio, ultima_fila)}"
    ws.print_title_rows = f"{fila_inicio}:{fila_inicio}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _pdf_estilos():
    styles = getSampleStyleSheet()
    return {
        "meta": ParagraphStyle(
            "MetaCharat", parent=styles["Normal"], fontSize=8,
            textColor=colors.HexColor(f"#{COLOR_TEXTO_SUAVE}"), leading=11,
        ),
        "celda": ParagraphStyle(
            "CeldaCharat", parent=styles["Normal"], fontSize=7,
            leading=9, textColor=colors.HexColor(f"#{COLOR_TEXTO}"),
        ),
        "kpi_lbl": ParagraphStyle(
            "KpiLbl", parent=styles["Normal"], fontSize=7,
            alignment=1, textColor=colors.white, fontName="Helvetica-Bold",
        ),
        "kpi_val": ParagraphStyle(
            "KpiVal", parent=styles["Normal"], fontSize=14,
            alignment=1, fontName="Helvetica-Bold", textColor=colors.HexColor(f"#{COLOR_TEXTO}"),
        ),
    }


def _pdf_tabla_kpi(reportes, resumen):
    est = _pdf_estilos()
    lbl_verde = ParagraphStyle("KpiLblV", parent=est["kpi_lbl"], textColor=colors.white)
    lbl_oscuro = ParagraphStyle("KpiLblO", parent=est["kpi_lbl"], textColor=colors.HexColor(f"#{COLOR_TEXTO}"))
    datos = [
        [Paragraph("TOTAL<br/>REGISTROS", lbl_verde), Paragraph("PENDIENTES", lbl_oscuro),
         Paragraph("EN PROCESO", lbl_oscuro), Paragraph("ATENDIDAS", lbl_oscuro)],
        [Paragraph(str(len(reportes)), ParagraphStyle("Kv0", parent=est["kpi_val"], textColor=colors.white)),
         Paragraph(str(resumen["Pendiente"]), est["kpi_val"]),
         Paragraph(str(resumen["En Proceso"]), est["kpi_val"]),
         Paragraph(str(resumen["Atendido"]), est["kpi_val"])],
    ]
    tabla = Table(datos, colWidths=[6.5 * cm, 6.5 * cm, 6.5 * cm, 6.5 * cm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 1), colors.HexColor(f"#{COLOR_VERDE}")),
        ("BACKGROUND", (1, 0), (1, 1), colors.HexColor("#FEF9C3")),
        ("BACKGROUND", (2, 0), (2, 1), colors.HexColor("#DBEAFE")),
        ("BACKGROUND", (3, 0), (3, 1), colors.HexColor("#DCFCE7")),
        ("TEXTCOLOR", (0, 0), (0, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor(f"#{COLOR_GRIS_BORDE}")),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor(f"#{COLOR_GRIS_BORDE}")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    return tabla


def _exportar_pdf(reportes, filtros=None):
    filtros = filtros or {}
    buf = io.BytesIO()
    generado = datetime.now().strftime("%d/%m/%Y %H:%M")
    resumen = _resumen_estados(reportes)
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.4 * cm, rightMargin=1.4 * cm,
        topMargin=3.1 * cm, bottomMargin=1.6 * cm,
    )
    est = _pdf_estilos()
    elementos = [
        Paragraph(
            f"<b>Período del reporte:</b> según filtros aplicados &nbsp;|&nbsp; "
            f"<b>Filtros:</b> {_etiqueta_filtros_export(filtros)} &nbsp;|&nbsp; "
            f"<b>Generado:</b> {generado} &nbsp;|&nbsp; "
            f"<b>Con GPS:</b> {sum(1 for r in reportes if _format_gps(r) != '—')}",
            est["meta"],
        ),
        Spacer(1, 0.25 * cm),
        _pdf_tabla_kpi(reportes, resumen),
        Spacer(1, 0.45 * cm),
    ]

    encabezados = COLUMNAS_EXPORT
    filas = [encabezados]
    for r in reportes:
        desc = (r.get("detalle", {}).get("descripcion") or "—").replace("&", "&amp;").replace("<", "&lt;")
        filas.append([
            r.get("gestion", {}).get("folio", "—"),
            _format_fecha(r.get("gestion", {}).get("fecha_registro")),
            r.get("informante", {}).get("nombre", ""),
            r.get("informante", {}).get("dni", ""),
            r.get("ubicacion", {}).get("caserio", ""),
            r.get("detalle", {}).get("categoria", ""),
            r.get("gestion", {}).get("estado", ""),
            _format_gps(r),
            Paragraph(desc, est["celda"]),
        ])

    if len(filas) == 1:
        filas.append(["Sin registros para los filtros seleccionados."] + [""] * (len(encabezados) - 1))

    col_widths = [2.8 * cm, 2.5 * cm, 2.8 * cm, 1.6 * cm, 2.5 * cm, 2.8 * cm, 2.0 * cm, 3.0 * cm, 6.5 * cm]
    tabla = Table(filas, colWidths=col_widths, repeatRows=1)
    estilos = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{COLOR_VERDE_OSC}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor(f"#{COLOR_GRIS_BORDE}")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{COLOR_GRIS}")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    for i, r in enumerate(reportes, start=1):
        est_row = r.get("gestion", {}).get("estado", "")
        color_est = {
            "Pendiente": colors.HexColor("#FEF9C3"),
            "En Proceso": colors.HexColor("#DBEAFE"),
            "Atendido": colors.HexColor("#DCFCE7"),
        }.get(est_row)
        if color_est:
            estilos.append(("BACKGROUND", (6, i), (6, i), color_est))
            estilos.append(("FONTNAME", (6, i), (6, i), "Helvetica-Bold"))
        estilos.append(("FONTNAME", (0, i), (0, i), "Helvetica-Bold"))
        estilos.append(("TEXTCOLOR", (0, i), (0, i), colors.HexColor(f"#{COLOR_VERDE_OSC}")))
    tabla.setStyle(TableStyle(estilos))

    elementos.append(tabla)
    doc._charat_generado = generado
    doc.build(elementos, onFirstPage=_pdf_pagina, onLaterPages=_pdf_pagina)
    buf.seek(0)
    return buf


def _pdf_pagina(canvas_obj, doc):
    w, h = doc.pagesize
    canvas_obj.saveState()
    for i, color in enumerate([COLOR_AMARILLO, COLOR_VERDE, COLOR_AZUL]):
        canvas_obj.setFillColor(colors.HexColor(f"#{color}"))
        canvas_obj.rect(0, h - (0.28 + i * 0.14) * cm, w, 0.12 * cm, fill=1, stroke=0)

    logo_x = doc.leftMargin
    logo_y = h - 2.55 * cm
    if os.path.isfile(LOGO_PATH):
        canvas_obj.drawImage(
            LOGO_PATH, logo_x, logo_y, width=1.55 * cm, height=1.55 * cm,
            preserveAspectRatio=True, mask="auto",
        )
        texto_x = logo_x + 1.85 * cm
    else:
        texto_x = logo_x

    canvas_obj.setFont("Helvetica-Bold", 13)
    canvas_obj.setFillColor(colors.HexColor(f"#{COLOR_VERDE_OSC}"))
    canvas_obj.drawString(texto_x, h - 1.55 * cm, "MUNICIPALIDAD DISTRITAL DE CHARAT")
    canvas_obj.setFont("Helvetica", 8)
    canvas_obj.setFillColor(colors.HexColor(f"#{COLOR_TEXTO_SUAVE}"))
    canvas_obj.drawString(texto_x, h - 1.9 * cm, "Provincia de Pataz — La Libertad, Perú")
    canvas_obj.drawString(texto_x, h - 2.2 * cm, "Reporte oficial de incidencias — Sistema MuniCharat")

    canvas_obj.setStrokeColor(colors.HexColor(f"#{COLOR_GRIS_BORDE}"))
    canvas_obj.setLineWidth(0.6)
    canvas_obj.line(doc.leftMargin, h - 2.75 * cm, w - doc.rightMargin, h - 2.75 * cm)

    canvas_obj.setStrokeColor(colors.HexColor(f"#{COLOR_GRIS_BORDE}"))
    canvas_obj.setLineWidth(0.4)
    canvas_obj.line(doc.leftMargin, 1.25 * cm, w - doc.rightMargin, 1.25 * cm)
    canvas_obj.setFont("Helvetica", 7)
    canvas_obj.setFillColor(colors.HexColor("#94A3B8"))
    canvas_obj.drawString(
        doc.leftMargin, 0.75 * cm,
        "Documento oficial generado por MuniCharat — Municipalidad Distrital de Charat",
    )
    fecha_doc = getattr(doc, "_charat_generado", datetime.now().strftime("%d/%m/%Y %H:%M"))
    canvas_obj.drawRightString(
        w - doc.rightMargin, 0.75 * cm,
        f"Página {canvas_obj.getPageNumber()} · {fecha_doc}",
    )
    canvas_obj.restoreState()


@app.get("/health")
def health():
    """Para comprobar que el proceso responde (Render / balanceadores)."""
    return "ok", 200


def optimizar_imagen(archivo_foto):
    img = Image.open(archivo_foto)
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    img.thumbnail((800, 800))
    buffer = io.BytesIO()
    img.save(buffer, format="JPEG", quality=60, optimize=True)
    buffer.seek(0)
    return buffer

# FUNCIÓN QUE CORRE EN SEGUNDO PLANO
def enviar_correo_async(app_context, categoria, nombre, caserio, descripcion):
    with app_context:
        try:
            config_doc = db.collection('configuracion').document('email_settings').get()
            if config_doc.exists:
                conf = config_doc.to_dict()
                correo_receptor = conf.get('correo_receptor')
                if correo_receptor:
                    asunto = f"MuniCharat: {categoria}"
                    cuerpo = f"Nuevo reporte de {nombre}.\nLugar: {caserio}\nDetalle: {descripcion}"
                    send_email(to_email=correo_receptor, subject=asunto, body_text=cuerpo)
                    print("LOG: Correo enviado exitosamente (Gmail API).")
        except Exception as e:
            print(f"LOG ERROR CORREO: {e}")

def _listar_sectores():
    try:
        docs = db.collection('sectores').order_by('nombre').stream()
        return [{"id": doc.id, "nombre": doc.to_dict().get('nombre', 'N/A'), "tipo": doc.to_dict().get('tipo', 'Sector')} for doc in docs]
    except Exception:
        return []

def _listar_categorias():
    try:
        docs = db.collection('categorias').order_by('nombre').stream()
        return [{"id": doc.id, "nombre": doc.to_dict().get('nombre', 'N/A'), "icono": _icono_categoria(doc.to_dict().get('nombre', ''))} for doc in docs]
    except Exception:
        return []

_sembrar_categorias_iniciales()

@app.route('/')
def index():
    return render_template('index.html', sectores=_listar_sectores(), categorias=_listar_categorias())

@app.route('/enviar', methods=['POST'])
def enviar():
    try:
        dni = request.form.get('dni')
        nombre = request.form.get('nombre')
        categoria = request.form.get('categoria')
        descripcion = request.form.get('descripcion')
        caserio = request.form.get('caserio')
        
        if not re.match(r"^\d{8}$", dni):
            flash("El DNI debe tener exactamente 8 dígitos numéricos.", "error")
            return redirect(url_for("index"))

        categorias_validas = {c['nombre'] for c in _listar_categorias()}
        if categoria not in categorias_validas:
            flash("Seleccione una categoría válida.", "error")
            return redirect(url_for("index"))

        sectores_validos = {s['nombre'] for s in _listar_sectores()}
        if caserio not in sectores_validos:
            flash("Seleccione un caserío o sector válido.", "error")
            return redirect(url_for("index"))

        lista_urls_fotos = []
        archivos = request.files.getlist('foto') 
        for archivo in archivos:
            if archivo and archivo.filename != '':
                nombre_unico = f"incidencias/{uuid.uuid4()}.jpg"
                blob = bucket.blob(nombre_unico)
                foto_comprimida = optimizar_imagen(archivo)
                blob.upload_from_file(foto_comprimida, content_type='image/jpeg')
                blob.make_public()
                lista_urls_fotos.append(blob.public_url)

        ahora = datetime.now()
        folio = _generar_folio()
        nueva_incidencia = {
            "informante": {"nombre": nombre, "dni": dni},
            "ubicacion": {"caserio": caserio, "lat": request.form.get('latitud'), "lng": request.form.get('longitud')},
            "detalle": {"categoria": categoria, "descripcion": descripcion, "fotos": lista_urls_fotos},
            "gestion": {
                "folio": folio,
                "estado": "Pendiente",
                "fecha_registro": ahora,
                "ultima_actualizacion": ahora,
                "historial": [{"estado": "Pendiente", "usuario": "Ciudadano", "fecha": ahora}],
                "comentarios": [],
            },
        }
        
        # GUARDADO EN FIREBASE
        db.collection('incidencias').add(nueva_incidencia)

        # DISPARAR CORREO EN HILO SEPARADO
        threading.Thread(target=enviar_correo_async, 
                         args=(app.app_context(), categoria, nombre, caserio, descripcion)).start()

        return render_template('exito.html', folio=folio)

    except Exception as e:
        print(f"ERROR CRÍTICO: {e}")
        flash("No se pudo registrar el reporte. Intente nuevamente.", "error")
        return redirect(url_for("index"))

@app.route('/consultar', methods=['GET', 'POST'])
def consultar():
    folio_buscado = None
    resultado = None
    if request.method == 'POST':
        folio_buscado = (request.form.get('folio') or '').strip().upper()
        if folio_buscado:
            docs = db.collection('incidencias').where('gestion.folio', '==', folio_buscado).limit(1).get()
            if docs:
                resultado = docs[0].to_dict()
    return render_template('consultar.html', folio_buscado=folio_buscado, resultado=resultado)

# --- RUTAS ADMINISTRATIVAS ---

@app.route('/admin')
def admin():
    if not session.get('logged_in'): return redirect(url_for('login'))
    reportes = _obtener_reportes()
    sectores = _listar_sectores()
    categorias = _listar_categorias()
    config_actual = {}
    usuarios = []
    if session.get('rol') == 'admin':
        usuarios = [doc.to_dict() | {"id": doc.id} for doc in db.collection('usuarios').stream()]
        conf_doc = db.collection('configuracion').document('email_settings').get()
        if conf_doc.exists: config_actual = conf_doc.to_dict()
    return render_template('admin.html', reportes=reportes, sectores=sectores, categorias=categorias,
                           usuarios=usuarios, config=config_actual, stats=_estadisticas_reportes(reportes),
                           analytics=_estadisticas_avanzadas(reportes), mapa_puntos=_puntos_mapa(reportes))

@app.route('/limpiar_datos_prueba', methods=['POST'])
def limpiar_datos_prueba():
    if not session.get('logged_in') or session.get('rol') != 'admin':
        return "No autorizado", 403
    if request.form.get("confirmacion", "").strip().upper() != "LIMPIAR":
        return _admin_redirect("Debe escribir LIMPIAR para confirmar la eliminación.", "warning")
    if not request.form.get("acepto"):
        return _admin_redirect("Debe aceptar que la acción es irreversible.", "warning")
    try:
        resultado = _limpiar_datos_prueba()
        mensaje = (
            f"Limpieza completada: {resultado['incidencias']} incidencia(s) y "
            f"{resultado['fotos']} foto(s) eliminadas. Usuarios, sectores y categorías se conservaron."
        )
        return _admin_redirect(mensaje)
    except Exception as e:
        print(f"ERROR LIMPIEZA: {e}")
        return _admin_redirect("No se pudo completar la limpieza. Revise los logs del servidor.", "error")


@app.route('/guardar_config_correo', methods=['POST'])
def guardar_config_correo():
    if not session.get('logged_in') or session.get('rol') != 'admin': return "No autorizado", 403
    db.collection('configuracion').document('email_settings').set({
        "correo_receptor": request.form.get('correo_receptor').strip(),
        "ultima_actualizacion": datetime.now()
    })
    return _admin_redirect("Correo de alertas actualizado correctamente.")

ESTADOS_INCIDENCIA = frozenset({"Pendiente", "En Proceso", "Atendido"})
ROLES_USUARIO = frozenset({"admin", "operador"})

@app.route('/cambiar_estado/<incidencia_id>/<path:estado>')
def cambiar_estado(incidencia_id, estado):
    if not session.get('logged_in'): return redirect(url_for('login'))
    if estado not in ESTADOS_INCIDENCIA: return "Estado no válido", 400
    ref = db.collection('incidencias').document(incidencia_id)
    if not ref.get().exists: return "Incidencia no encontrada", 404
    _registrar_historial(ref, estado, session.get('user', 'Operador'))
    return _admin_redirect(f"Incidencia marcada como «{estado}».")

@app.route('/agregar_comentario/<incidencia_id>', methods=['POST'])
def agregar_comentario(incidencia_id):
    if not session.get('logged_in'): return redirect(url_for('login'))
    texto = (request.form.get('comentario') or '').strip()
    if not texto: return _admin_redirect("Escriba un comentario.", "warning")
    ref = db.collection('incidencias').document(incidencia_id)
    if not ref.get().exists: return _admin_redirect("Incidencia no encontrada.", "error")
    ref.update({
        "gestion.comentarios": ArrayUnion([{
            "texto": texto,
            "usuario": session.get('user', 'Operador'),
            "fecha": datetime.now(),
        }]),
        "gestion.ultima_actualizacion": datetime.now(),
    })
    return _admin_redirect("Comentario interno agregado.")

@app.route('/exportar/excel')
def exportar_excel():
    if not session.get('logged_in'): return redirect(url_for('login'))
    reportes = _filtrar_reportes(_obtener_reportes(), request.args)
    buf = _exportar_excel(reportes, request.args)
    nombre = f"incidencias_charat_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nombre, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/exportar/pdf')
def exportar_pdf():
    if not session.get('logged_in'): return redirect(url_for('login'))
    reportes = _filtrar_reportes(_obtener_reportes(), request.args)
    buf = _exportar_pdf(reportes, request.args)
    nombre = f"incidencias_charat_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(buf, as_attachment=True, download_name=nombre, mimetype="application/pdf")

@app.route('/agregar_sector', methods=['POST'])
def agregar_sector():
    if not session.get('logged_in') or session.get('rol') != 'admin': return "No autorizado", 403
    nombre = (request.form.get('nuevo_sector') or '').strip()
    if not nombre: return _admin_redirect("Ingrese un nombre de sector.", "warning")
    db.collection('sectores').add({
        "nombre": nombre,
        "tipo": request.form.get('tipo_sector', 'Sector'),
        "fecha_registro": datetime.now()
    })
    return _admin_redirect(f"Sector «{nombre}» registrado correctamente.")

@app.route('/eliminar_sector/<sector_id>')
def eliminar_sector(sector_id):
    if not session.get('logged_in') or session.get('rol') != 'admin': return "No autorizado", 403
    db.collection('sectores').document(sector_id).delete()
    return _admin_redirect("Sector eliminado correctamente.")

@app.route('/agregar_categoria', methods=['POST'])
def agregar_categoria():
    if not session.get('logged_in') or session.get('rol') != 'admin': return "No autorizado", 403
    nombre = (request.form.get('nueva_categoria') or '').strip()
    if not nombre: return _admin_redirect("Ingrese un nombre de categoría.", "warning")
    db.collection('categorias').add({
        "nombre": nombre,
        "fecha_registro": datetime.now()
    })
    return _admin_redirect(f"Categoría «{nombre}» registrada correctamente.")

@app.route('/eliminar_categoria/<categoria_id>')
def eliminar_categoria(categoria_id):
    if not session.get('logged_in') or session.get('rol') != 'admin': return "No autorizado", 403
    db.collection('categorias').document(categoria_id).delete()
    return _admin_redirect("Categoría eliminada correctamente.")

@app.route('/crear_usuario', methods=['POST'])
def crear_usuario():
    if not session.get('logged_in') or session.get('rol') != 'admin': return "No autorizado", 403
    usuario = (request.form.get('nuevo_usuario') or '').strip()
    password = request.form.get('nueva_password') or ''
    rol = request.form.get('rol', 'operador')
    if not usuario or not password: return _admin_redirect("Complete usuario y contraseña.", "warning")
    if rol not in ROLES_USUARIO: rol = 'operador'
    existente = db.collection('usuarios').where('usuario', '==', usuario).limit(1).get()
    if existente: return _admin_redirect("Ese usuario ya existe.", "warning")
    db.collection('usuarios').add({
        "usuario": usuario,
        "password": generate_password_hash(password),
        "rol": rol,
        "fecha_registro": datetime.now()
    })
    return _admin_redirect(f"Usuario «{usuario}» creado correctamente.")

@app.route('/eliminar_usuario/<usuario_id>')
def eliminar_usuario(usuario_id):
    if not session.get('logged_in') or session.get('rol') != 'admin': return "No autorizado", 403
    ref = db.collection('usuarios').document(usuario_id)
    doc = ref.get()
    if not doc.exists: return _admin_redirect("Usuario no encontrado.", "warning")
    if doc.to_dict().get('usuario') == session.get('user'): return _admin_redirect("No puedes eliminar tu propia cuenta.", "error")
    ref.delete()
    return _admin_redirect("Usuario eliminado correctamente.")

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        u, p = request.form.get('usuario'), request.form.get('password')
        docs = db.collection('usuarios').where('usuario', '==', u).limit(1).get()
        if docs:
            doc = docs[0]
            data = doc.to_dict()
            if _verificar_password(data.get('password', ''), p):
                if not _password_hasheada(data.get('password', '')):
                    doc.reference.update({'password': generate_password_hash(p)})
                session.update({'logged_in': True, 'user': data['usuario'], 'rol': data.get('rol', 'operador')})
                return redirect(url_for('admin'))
        return render_template('login.html', error="Credenciales incorrectas")
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)